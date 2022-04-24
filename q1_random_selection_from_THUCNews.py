# import imp
import os
import random
# import re
# import string
# import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

rootpath='I:/THUCNews/'
raw_data_true=[]
raw_data_false=[]
sheetpath='D:/Coding-Projects/Data_Cleaning_2022_Teddy_Cup_C/data_THUCNews.xlsx'

def sub_obj(path):
    sub_obj=[]
    # test_i=0
    for i in os.listdir(path):
        # test_i+=1
        # if test_i%5000==0:
        #     print(i)
        sub_obj.append(i)
    return sub_obj


if __name__=='__main__':
    sub_class=sub_obj(rootpath)
    contained=[]
    sub_files=[]
    # test_i=0
    for i in sub_class:
        # test_i+=1
        # if test_i%10000==0:
        #     print(test_i)
        subpath=rootpath+i+'/'
        sub_files.append(sub_obj(subpath))
    j=0
    while(j<6000):                                              #Total number of files
        if (j+1)%10==0:
            print(j+1)
        folders_count=len(sub_class)
        # sub_files=sub_obj(i)
        # files_count=len(sub_files)
        class_op=random.randint(0,folders_count-1)
        files_count=len(sub_files[class_op])
        file_op=random.randint(0,files_count-1)
        if str(class_op)+str(file_op) not in contained:
            with open(rootpath+sub_class[class_op]+'/'+sub_files[class_op][file_op],'r',encoding='UTF-8') as text:
                # raw_line=text.read()
                raw_data_false.append(text.read().replace('\n','').replace('　','')+'\n')
            contained.append(str(class_op)+'/'+str(file_op))
            text.close()
            j+=1

    wb=Workbook()
    sheet=wb.active
    for i in range(6000):
        sheet.cell(row=i+1,column=1).value=ILLEGAL_CHARACTERS_RE.sub(r'',raw_data_false[i])
        sheet.cell(row=i+1,column=2).value='false'
    wb.save(sheetpath)

    # with open('F:/THUCNews/collection_12k_false.txt','w',encoding='UTF-8') as data:
    #     data.writelines(raw_data_false)
    # data.close()
    