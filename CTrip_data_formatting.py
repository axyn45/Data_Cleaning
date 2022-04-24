# from curses import raw
# import os
# import random
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


data_list=[]
raw_sheet_path='D:/Coding-Projects/Data_Cleaning_2022_Teddy_Cup_C/CTrip_raw_data.xlsx'
sheetpath='D:/Coding-Projects/Data_Cleaning_2022_Teddy_Cup_C/true_data_CTrip_3.6k.xlsx'

if __name__=='__main__':
    raw_sheet=load_workbook(raw_sheet_path)
    raw_sheet.active
    data=Workbook()
    data_sheet=data.active

    for i in range(1,3600):
        data_list.append((raw_sheet.active.cell(row=i,column=1).value)+'#'+(raw_sheet.active.cell(row=i,column=2).value).replace('…',''))

    for i in range(1,3600):
        data_sheet.cell(row=i,column=1).value=ILLEGAL_CHARACTERS_RE.sub(r'',data_list[i-1])
        data_sheet.cell(row=i,column=2).value='true'  

    data.save(sheetpath)
